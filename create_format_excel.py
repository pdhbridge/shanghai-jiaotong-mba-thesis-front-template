import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print('请先安装openpyxl库: pip install openpyxl')
    import sys
    sys.exit(1)

# 创建工作簿
wb = openpyxl.Workbook()

# 定义样式
header_font = Font(name='SimHei', size=12, bold=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== 第1个Sheet：页面设置 ====================
ws1 = wb.active
ws1.title = "页面设置"

# 设置列宽
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 50

# 标题行
ws1['A1'] = '设置项目'
ws1['B1'] = '具体参数'
ws1['C1'] = '备注说明'

for col in ['A1', 'B1', 'C1']:
    ws1[col].font = header_font
    ws1[col].fill = header_fill
    ws1[col].alignment = header_alignment
    ws1[col].border = thin_border

# 页面设置数据
page_settings = [
    ['纸张规格', 'A4', '标准A4纸张'],
    ['页眉距离', '2.5 cm', '从页面顶部到页眉的距离'],
    ['页脚距离', '3.0 cm', '从页面底部到页脚的距离'],
    ['上边距', '3.5 cm', '页面顶部边距'],
    ['下边距', '4.0 cm', '页面底部边距'],
    ['左边距', '2.8 cm', '页面左侧边距'],
    ['右边距', '2.8 cm', '页面右侧边距'],
    ['打印方式', '计算机编辑输出', '论文内容一律采用计算机编辑'],
    ['装订方式', '研究生院统一封面线装成册', '使用学校统一封面'],
    ['特殊情况', '论文幅面允许根据实际需要延长和加宽', '如图样、表不能缩小时'],
]

for row_idx, row_data in enumerate(page_settings, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

# ==================== 第2个Sheet：字体样式要求 ====================
ws2 = wb.create_sheet(title="字体样式要求")

# 设置列宽
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 40

# 标题行
headers2 = ['应用位置', '字体名称', '字号', '加粗', '对齐方式', '备注说明']
for col_idx, header in enumerate(headers2, start=1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 字体样式数据
font_styles = [
    ['扉页-中文题目', '黑体', '二号', '否', '居中', '可分成1-3行居中打印'],
    ['扉页-英文题目', '二号字体', '二号', '否', '居中', '全部大写，可分成1-3行，每行左右两边至少留五个字符空格'],
    ['扉页-学校、学院、专业、作者、导师、学号、班级及答辩日期', '黑体', '三号', '否', '居中', '均独立成行'],
    ['页眉左边', '宋体', '小五号', '否', '左对齐', '"上海交通大学MBA学位论文"'],
    ['页眉右边', '宋体', '小五号', '否', '右对齐', '论文的题目'],
    ['页脚', '-', '-', '否', '居中', '页码'],
    ['中文摘要-论文题目', '黑体', '三号', '否', '居中', '可以分成1或2行居中打印'],
    ['中文摘要-"摘要"二字', '黑体', '三号', '否', '居中', '字间空一格'],
    ['中文摘要-摘要内容', '宋体', '四号', '否', '左对齐', '每段开头空二格'],
    ['中文摘要-"关键词"三字', '黑体', '四号', '否', '左对齐', '-'],
    ['中文摘要-关键词', '宋体', '四号', '否', '左对齐', '数量4~6个，每一关键词之间用逗号分开'],
    ['英文摘要-英文题目', '三号字体', '三号', '否', '居中', '全部大写，可分成1~3行，每行左右两边至少留五个字符空格'],
    ['英文摘要-"ABSTRACT"', '三号字体', '三号', '否', '居中', '-'],
    ['英文摘要-摘要内容', '四号字体', '四号', '否', '左对齐', '每段开头留四个字符空格'],
    ['英文摘要-"KEY WORDS"', '四号字体', '四号', '否', '左对齐', '关键词小写'],
    ['英文摘要-关键词', '四号字体', '四号', '否', '左对齐', '每一关键词之间用逗号分开'],
    ['目录-"目录"两字', '黑体', '三号', '否', '居中', '-'],
    ['目录-内容文字', '宋体', '五号', '否', '左对齐', '章、节、小节分别以"第1章"、"1.1"、"1.1.1"等标出'],
    ['章标题', '黑体', '三号', '否', '居中', '每章标题以三号黑体居中打印'],
    ['节标题（如"1.1"）', '黑体', '四号', '否', '左对齐（顶格）', '"章"下空两行为节'],
    ['小节标题（如"1.1.1"）', '黑体', '小四号/五号', '否', '左对齐（顶格）', '节下空一行为小节'],
    ['正文', '宋体', '五号/小四号', '否', '左对齐', '-'],
    ['图题-中文', '楷体', '五号', '否', '居中', '写在图的正下方居中位置'],
    ['图题-英文', '五号字体', '五号', '否', '居中', '图题采用中英文对照'],
    ['图-文献来源', '宋体', '小五号', '否', '左对齐', '引用图应在图题左下角标出文献来源'],
    ['表题', '楷体', '五号', '否', '居中', '写在表格的正上方居中'],
    ['参考文献标题', '黑体', '三号', '否', '居中', '"参考文献"四字'],
    ['参考文献内容', '宋体', '五号', '否', '左对齐', '-'],
    ['附录编号和标题', '黑体', '四号', '否', '居中', '-'],
    ['附录内容', '宋体', '五号', '否', '左对齐', '-'],
    ['致谢标题"致谢"', '黑体', '三号', '否', '居中', '字间空一格'],
    ['致谢内容', '宋体', '小四号', '否', '左对齐', '-'],
]

for row_idx, row_data in enumerate(font_styles, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

# ==================== 第3个Sheet：段落格式要求 ====================
ws3 = wb.create_sheet(title="段落格式要求")

# 设置列宽
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 50

# 标题行
headers3 = ['应用位置', '格式参数', '具体要求']
for col_idx, header in enumerate(headers3, start=1):
    cell = ws3.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 段落格式数据
paragraph_formats = [
    ['章标题', '占行', '占两行'],
    ['章、条的编号', '排版位置', '顶格排'],
    ['章、条的编号与标题/文字', '间距', '一个字的间隙'],
    ['段的文字', '起排位置', '空两个字起排'],
    ['段的文字回行', '排版方式', '顶格排'],
    ['正文总项中的分项', '序号格式', '（1）、（2）、（3）…'],
    ['分项中的小项', '序号格式', '①、②、③…'],
    ['分项序号书写位置', '起排位置', '均从顶格起空两格书写'],
    ['扉页打印顺序-首行', '空行', '从首行下空3-4行'],
    ['扉页打印顺序-中文题目后', '空行', '下空1行'],
    ['扉页打印顺序-英文题目后', '空行', '下空8-11行（可自行调整）'],
    ['中文摘要-论文题目后', '空行', '下空一行'],
    ['中文摘要-"摘要"二字后', '空行', '下空一行'],
    ['英文摘要-题目后', '空行', '下空三行'],
    ['英文摘要-"ABSTRACT"后', '空行', '下空二行'],
    ['英文摘要-摘要内容后', '空行', '下空二行'],
    ['目录-"目录"两字后', '空行', '下空两行'],
    ['章标题后', '空行', '"章"下空两行为节'],
    ['节标题后', '空行', '节下空一行为小节'],
    ['参考文献标题后', '空行', '下空一行'],
    ['致谢标题后', '空行', '下空一行'],
    ['附录', '换页方式', '每一个附录应另起一面，以后各个附录通常另起一面'],
    ['中文摘要、英文摘要、目录、前言以及每一章的开始', '换页方式', '均须从独立一页开始打印'],
    ['参考文献', '换页方式', '另起一页打印'],
]

for row_idx, row_data in enumerate(paragraph_formats, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

# ==================== 第4个Sheet：编号格式要求 ====================
ws4 = wb.create_sheet(title="编号格式要求")

# 设置列宽
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 50

# 标题行
headers4 = ['编号类型', '编号格式', '具体说明']
for col_idx, header in enumerate(headers4, start=1):
    cell = ws4.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 编号格式数据
numbering_formats = [
    ['章标题（一级）', '第1章、第2章、第3章', '各级末尾不加标点，编号与后面的文字之间应有1个字的间隔'],
    ['节标题（二级）', '1.1、1.2、1.3', '各级末尾不加标点，编号与后面的文字之间应有1个字的间隔'],
    ['小节标题（三级）', '1.1.1、1.1.2、1.1.3', '各级末尾不加标点，编号与后面的文字之间应有1个字的间隔'],
    ['图编号', '图1、图2、图3', '由"图"和从1开始的阿拉伯数字组成，编号一直连续到附录之前，与章、条和表的编号无关'],
    ['表编号', '表1、表2、表3', '由"表"和从1开始的阿拉伯数字组成，编号一直连续到附录之前，与章、条和图的编号无关'],
    ['公式编号', '(1-1)、(1-2)', '公式后应注明编号，用省略号连接，按章顺序编排'],
    ['参考文献序号', '[1]、[2]、[3]', '按论文中参考文献出现的先后顺序用阿拉伯数字连续编号，将序号置于方括号内'],
    ['附录编号', '附录1、附录2', '依次为附录1，附录2，……'],
    ['续表编号', '表1（续）', '如某个表需要转页接排，在随后的各页上应重复表的编号，编号后跟表题（可省略）和"（续）"'],
    ['正文中分项序号', '（1）、（2）、（3）', '正文中总项包括的分项'],
    ['分项中小项序号', '①、②、③', '对分项中的小项，均从顶格起空两格书写'],
    ['目录章编号', '第1章', '章、节、小节分别以"第1章"、"1.1"、"1.1.1"等数字依次标出'],
    ['目录节编号', '1.1', '目录中章、条的编号和绪论、附录等均顶格排'],
    ['目录小节编号', '1.1.1', '章、条的标题及附录等的标题与前面的内容之间空一个字的间隙'],
]

for row_idx, row_data in enumerate(numbering_formats, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

# ==================== 第5个Sheet：各部分格式要求 ====================
ws5 = wb.create_sheet(title="各部分格式要求")

# 设置列宽
ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 50
ws5.column_dimensions['C'].width = 50

# 标题行
headers5 = ['论文部分', '格式要求', '备注说明']
for col_idx, header in enumerate(headers5, start=1):
    cell = ws5.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 各部分格式数据
sections_formats = [
    ['封面（外封面、扉页）', '外封面由论文装订处提供，打印论文题目（不超过20个字）、学科专业、作者姓名、指导教师、答辩日期等；扉页需自行打印', '论文题目应避免使用不常用缩略语、首字母缩写字、字符、代号和公式等'],
    ['扉页', '中文题目：二号黑体居中；英文题目：二号字体，全部大写，居中；学校、学院、专业、作者、导师、学号、班级及答辩日期：三号黑体居中', 'A4纸，页眉2.5cm，页脚3.0cm，页边距上3.5cm，下4.0cm，左2.8cm，右2.8cm'],
    ['论文原创性声明', '全文可从网上下载，经学位论文作者签名后生效', '网址：www.gs.sjtu.edu.cn'],
    ['论文版权使用授权书', '全文可从网上下载，经学位论文作者和指导教师共同签名后生效', '网址：www.gs.sjtu.edu.cn'],
    ['论文答辩委员会名单及答辩决议', '须经全体答辩委员、答辩委员会主席签名，可以用复印件附于归档论文中', '学生可以忽略此条'],
    ['中文摘要', '论文题目：三号黑体居中；"摘要"二字：三号黑体居中，字间空一格；摘要内容：四号宋体，每段开头空二格；"关键词"：四号黑体；关键词：四号宋体', '字数一般不宜超过500字，关键词数量4~6个，用逗号分开'],
    ['英文摘要', '英文题目：三号字体，全部大写，居中；"ABSTRACT"：三号字体，居中；摘要内容：四号字体，每段开头留四个字符空格；"KEY WORDS"：四号字体；关键词：四号字体，小写', '内容应与中文摘要基本相对应，字数一般不宜超过300个实词'],
    ['目录', '"目录"两字：三号黑体居中；目录内容：五号宋体', '章、节、小节分别以"第1章"、"1.1"、"1.1.1"等数字依次标出，前言、章、条等与页码之间用"……"连接'],
    ['符号说明', '论文中所用符号所表示的意义及单位（或量纲）', '符号、标志、缩略词、首字母缩写、计量单位、名词、术语等注释说明'],
    ['论文正文', '章标题：三号黑体居中；节标题：四号黑体顶格；小节标题：小四号或五号黑体顶格；正文：五号或小四号宋体', '一般由标题、文字叙述、图、表格和公式等五个部分构成'],
    ['图', '图编号：图1、图2等；图题：五号楷体居中，写在图的正下方；图题采用中英文对照；文献来源：小五号宋体', '图中一律采用英文字母标注，图文说明用中文'],
    ['表格', '表编号：表1、表2等；表题：五号楷体居中，写在表格的正上方', '表内必须按规定的符号标注单位，续表应重复表头和关于单位的陈述'],
    ['公式', '公式应另起一行居中排；公式后应注明编号，用省略号连接，按章顺序编排', '公式下面的"式中："空两个字起排，单独占一行'],
    ['参考文献', '"参考文献"四字：三号黑体居中；参考文献各项目：五号宋体', '按论文中参考文献出现的先后顺序用阿拉伯数字连续编号，将序号置于方括号内，另起一页打印'],
    ['注释', '可作为脚注在页下分散著录', '切忌在文中注释'],
    ['附录', '附录编号和标题：四号黑体居中；附录内容：五号宋体', '每一个附录应另起一面，附录中的图表公式另编排序号，与正文分开'],
    ['致谢', '"致谢"：三号黑体居中，字间空一格；致谢内容：小四号宋体', '一般在正文之后，应实事求是，切忌浮夸与庸俗之词'],
]

for row_idx, row_data in enumerate(sections_formats, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

# ==================== 第6个Sheet：论文结构组成 ====================
ws6 = wb.create_sheet(title="论文结构组成")

# 设置列宽
ws6.column_dimensions['A'].width = 10
ws6.column_dimensions['B'].width = 30
ws6.column_dimensions['C'].width = 50

# 标题行
headers6 = ['序号', '组成部分', '说明']
for col_idx, header in enumerate(headers6, start=1):
    cell = ws6.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 论文结构数据
structure_data = [
    ['1', '封面（包括外封面、扉页）', '外封面由论文装订处提供，扉页需自行打印'],
    ['2', '题名页', 'MBA学位论文无需此页'],
    ['3', '论文原创性声明', '从网上下载，经学位论文作者签名后生效'],
    ['4', '论文版权使用授权书', '从网上下载，经学位论文作者和指导教师共同签名后生效'],
    ['5', '论文答辩委员会名单及答辩决议', '学生可以忽略此条'],
    ['6', '中文摘要（包括关键词）', '字数一般不宜超过500字'],
    ['7', '英文摘要（包括关键词）', '内容应与中文摘要基本相对应，字数一般不宜超过300个实词'],
    ['8', '目录', '将文内的章节标题依次排列'],
    ['9', '符号说明', '论文中所用符号所表示的意义及单位（或量纲）'],
    ['10', '论文正文（包括前言）', '是学位论文的核心部分'],
    ['11', '参考文献', '按文中引用出现的顺序列出，按照GB 7714规定执行'],
    ['12', '注释', '可作为脚注在页下分散著录'],
    ['13', '附录', '作为论文主体的补充项目，并不是必须的'],
    ['14', '致谢', '一般在正文之后'],
]

for row_idx, row_data in enumerate(structure_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

# 保存文件
output_path = os.path.join(os.path.dirname(__file__), '上海交通大学MBA学位论文格式要求.xlsx')
wb.save(output_path)
print(f'Excel文件已生成: {output_path}')